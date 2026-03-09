# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# exporter.py
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
"""Genera reportes en Excel (.xlsx) y PDF a partir de una lista de órdenes."""

import os
from datetime import date, datetime

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    EXCEL_OK = True
except ImportError:
    EXCEL_OK = False

try:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    PDF_OK = True
except ImportError:
    PDF_OK = False


class Exporter:
    """
    Responsabilidad única: serializar órdenes a Excel o PDF.
    Devuelve (ok: bool, mensaje: str) para que la UI decida cómo notificar.
    """

    HEADERS = ["ID", "Cliente", "Producto", "Fecha", "Monto", "Estado"]
    ESTADO_COLORS_XL = {
        "Completado": "D1FAE5",
        "Pendiente":  "FEF3C7",
        "Cancelado":  "FEE2E2",
    }

    # ── Excel ─────────────────────────────────────────────────────────────────
    def to_excel(self, orders: list[dict]) -> tuple[bool, str]:
        if not EXCEL_OK:
            return False, "Instala openpyxl:\npip install openpyxl"
        if not orders:
            return False, "No hay órdenes para exportar."

        fname = f"reporte_{date.today()}.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Órdenes"

        hfill = PatternFill("solid", fgColor="075E54")
        for col, h in enumerate(self.HEADERS, 1):
            c = ws.cell(row=1, column=col, value=h)
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = hfill
            c.alignment = Alignment(horizontal="center")
            ws.column_dimensions[c.column_letter].width = 20

        for row, o in enumerate(orders, 2):
            vals = [o["id"], o["cliente"], o["producto"],
                    o["fecha"], o["monto"], o["estado"]]
            fill = PatternFill("solid", fgColor=self.ESTADO_COLORS_XL.get(o["estado"], "FFFFFF"))
            for col, v in enumerate(vals, 1):
                c = ws.cell(row=row, column=col, value=v)
                c.fill = fill
                c.alignment = Alignment(horizontal="center")

        last = len(orders) + 2
        ws.cell(row=last, column=4, value="TOTAL").font = Font(bold=True)
        ws.cell(row=last, column=5, value=sum(o["monto"] for o in orders)).font = Font(bold=True)

        wb.save(fname)
        self._open_file(fname)
        return True, f"Archivo guardado:\n{fname}"

    # ── PDF ───────────────────────────────────────────────────────────────────
    def to_pdf(self, orders: list[dict], date_from: str, date_to: str) -> tuple[bool, str]:
        if not PDF_OK:
            return False, "Instala reportlab:\npip install reportlab"
        if not orders:
            return False, "No hay órdenes para exportar."

        fname  = f"reporte_{date.today()}.pdf"
        doc    = SimpleDocTemplate(fname, pagesize=A4)
        styles = getSampleStyleSheet()
        elems  = []

        elems.append(Paragraph("<b>Reporte de Órdenes — WhatsApp Business</b>", styles["Title"]))
        elems.append(Spacer(1, 6))
        elems.append(Paragraph(
            f"Período: {date_from} → {date_to}  |  "
            f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}",
            styles["Normal"]))
        elems.append(Spacer(1, 16))

        data = [self.HEADERS]
        for o in orders:
            data.append([o["id"], o["cliente"], o["producto"],
                         o["fecha"], f"${o['monto']:.2f}", o["estado"]])
        data.append(["", "", "", "TOTAL", f"${sum(o['monto'] for o in orders):.2f}", ""])

        tbl = Table(data, colWidths=[60, 110, 110, 70, 65, 80])
        tbl.setStyle(TableStyle([
            ("BACKGROUND",     (0, 0),  (-1, 0),  colors.HexColor("#075E54")),
            ("TEXTCOLOR",      (0, 0),  (-1, 0),  colors.white),
            ("FONTNAME",       (0, 0),  (-1, 0),  "Helvetica-Bold"),
            ("FONTSIZE",       (0, 0),  (-1, 0),  10),
            ("ALIGN",          (0, 0),  (-1, -1), "CENTER"),
            ("ROWBACKGROUNDS", (0, 1),  (-1, -2), [colors.white, colors.HexColor("#F0FDF4")]),
            ("FONTNAME",       (0, -1), (-1, -1), "Helvetica-Bold"),
            ("LINEBELOW",      (0, 0),  (-1, 0),  1, colors.white),
            ("GRID",           (0, 0),  (-1, -1), 0.4, colors.HexColor("#E5E7EB")),
        ]))
        elems.append(tbl)

        doc.build(elems)
        self._open_file(fname)
        return True, f"Archivo guardado:\n{fname}"

    # ── Helper ────────────────────────────────────────────────────────────────
    @staticmethod
    def _open_file(path: str):
        if os.name == "nt":
            os.startfile(path)
        else:
            os.system(f'xdg-open "{path}" 2>/dev/null || open "{path}" 2>/dev/null &')

